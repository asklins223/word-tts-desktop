"""「词汇」题型切片：Excel 单词导入模板解析器与题型元数据。"""

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from question_types.base import BaseParser
from question_types.text_utils import sanitize


# ============================================================================
# 7. 词汇解析器（Excel 单词导入模板）
# ============================================================================

class ExcelVocabularyParser(BaseParser):
    """
    解析 Excel 单词导入模板（.xlsx），提取「单词名称」和「例句」两列。

    每个单词生成两条 TTS 条目：
      1. 单词本身 — 使用默认女声 Amanda，命名「单词1」「单词2」…
      2. 例句     — 使用默认女声 Amanda，命名「句子1」「句子2」…

    Excel 结构（第一行为表头）：
      A: 单元    (如 Unit 6)
      B: 课时信息 (如 Understanding ideas)
      C: 单词名称 (如 pigeon)
      D: 美式音标
      E: 英式音标
      F: 词性
      G: 释义
      H: 例句    (如 A pigeon is standing near the window.)
      I: 翻译

    解析器自动识别「单词名称」和「例句」列的位置（按表头匹配），
    无需硬编码列号。
    """

    DOC_TYPE = "词汇"

    # Excel 文件不使用 load_paragraphs
    _SKIP_LOAD_PARAGRAPHS = True

    # 表头匹配关键词
    WORD_HEADER_KEYWORDS = ("单词名称", "单词")
    SENTENCE_HEADER_KEYWORDS = ("例句",)

    @staticmethod
    def _find_header_column(headers, keywords):
        """Prefer an exact template header before falling back to aliases."""

        normalized = {
            index: str(value).strip()
            for index, value in enumerate(headers, start=1)
            if value is not None and str(value).strip()
        }
        for keyword in keywords:
            for index, header in normalized.items():
                if header == keyword:
                    return index
        for keyword in keywords:
            for index, header in normalized.items():
                if keyword in header:
                    return index
        return None

    def parse(self):
        if not _OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "解析 Excel 文件需要 openpyxl 库，请运行: pip install openpyxl"
            )

        wb = None
        try:
            wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
            worksheets = []
            saw_word_header = False
            saw_sentence_header = False
            for ws in wb.worksheets:
                header_values = []
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    header_values = list(row)
                    break
                word_col = self._find_header_column(header_values, self.WORD_HEADER_KEYWORDS)
                sentence_col = self._find_header_column(header_values, self.SENTENCE_HEADER_KEYWORDS)
                saw_word_header = saw_word_header or word_col is not None
                saw_sentence_header = saw_sentence_header or sentence_col is not None
                # Excel templates often contain empty/instruction sheets.  Only
                # sheets with both required columns contribute vocabulary items.
                if word_col is not None and sentence_col is not None:
                    worksheets.append((ws, word_col, sentence_col))

            if not worksheets:
                if not saw_word_header:
                    raise ValueError(
                        "未找到「单词名称」列，请确认 Excel 表头包含「单词名称」或「单词」"
                    )
                if not saw_sentence_header:
                    raise ValueError(
                        "未找到「例句」列，请确认 Excel 表头包含「例句」"
                    )
                raise ValueError("未找到同时包含「单词名称」和「例句」列的工作表")

            # ---- 提取所有符合模板的工作表数据 ----
            items = []
            word_seq = 0       # 单词序号（跨工作表连续）
            sentence_seq = 0   # 例句序号（跨工作表连续）

            for ws, word_col, sentence_col in worksheets:
                for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    # openpyxl read_only 模式返回的行可能短于总列数
                    word_val = row[word_col - 1] if len(row) >= word_col else None
                    sentence_val = row[sentence_col - 1] if len(row) >= sentence_col else None

                    word_text = sanitize(str(word_val).strip()) if word_val else ""
                    sentence_text = sanitize(str(sentence_val).strip()) if sentence_val else ""

                    # 跳过空行
                    if not word_text and not sentence_text:
                        continue

                    # 单词条目：使用统一默认女声 Amanda，不设置单词专用音色。
                    if word_text:
                        word_seq += 1
                        items.append({
                            "category": "单词",
                            "number": word_seq,
                            "filename_stem": f"单词{word_seq}",
                            "voice": "female",
                            "text": word_text,
                            "source_locator": f"工作表/{ws.title}/行/{excel_row}/单词",
                        })

                    # 例句条目：同样使用统一默认女声 Amanda。
                    if sentence_text:
                        sentence_seq += 1
                        items.append({
                            "category": "例句",
                            "number": sentence_seq,
                            "filename_stem": f"句子{sentence_seq}",
                            "voice": "female",
                            "text": sentence_text,
                            "source_locator": f"工作表/{ws.title}/行/{excel_row}/例句",
                        })

            return self._result(items)
        finally:
            if wb is not None:
                wb.close()
