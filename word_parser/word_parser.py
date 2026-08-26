#!/usr/bin/env python3
"""
Word 文档解析脚本
=================
从 examples/documents 文件夹下的 Word 文档中提取各类英语听说考试素材。

支持文档类型：
  1. 信息获取        — 提取「听选信息」「回答问题」的题目与录音稿
  2. 听后选择        — 仅提取「录音原文」中的 W/M 对话
  3. 听后应答        — 提取计算机提示后的待朗读句子，按行生成
  4. 课文跟读        — 提取句子跟读（去序号、按序号排序）、段落跟读、语篇跟读
  5. 信息转述及询问   — 提取「第一节 信息转述」的录音稿
  6. 模仿朗读        — 提取每个单元的「外网」(2篇) 和「教材」(1篇)
  7. 词汇            — 预留接口，未来接入

设计说明：
  - 每种文档类型对应一个 Parser 子类，通过文件名自动识别
  - 使用状态机 + 循环遍历段落，天然支持同一文件内包含多篇试题
  - 循环终止条件为段落列表遍历完毕，辅以章节标记进行状态切换
  - 输出结构化 JSON，方便后续 TTS 处理

用法:
    python word_parser/word_parser.py

输出:
    examples/parsed/parsed_results.json
"""

import os
import re
import json
from docx import Document

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

# ============================================================================
# 路径配置（仅供解析器命令行示例使用）
# ============================================================================
PARSER_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(PARSER_DIR)
WORD_DIR = os.path.join(PROJECT_ROOT, "examples", "documents")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "examples", "parsed")


# ============================================================================
# 工具函数
# ============================================================================

def load_paragraphs(filepath):
    """
    加载 Word 文档，返回非空段落列表。
    每个元素为 (原始索引, 文本, 样式名)。
    """
    doc = Document(filepath)
    result = []
    for i, para in enumerate(doc.paragraphs):
        text = para.text.strip()
        if text:
            style = para.style.name if para.style else ""
            result.append((i, text, style))
    return result


def is_chinese(text):
    """判断文本是否以中文为主（CJK 字符占比 > 30%）"""
    cjk = sum(1 for c in text if '\u4e00' <= c <= '\u9fff')
    total = len(text.replace(' ', '').replace('\n', ''))
    return total > 0 and cjk / total > 0.3


def clean_whitespace(text):
    """规范化空白：合并连续空格，去除行首尾空格，保留换行"""
    lines = text.split('\n')
    cleaned = [re.sub(r'[ \t\u00a0]+', ' ', line.strip()) for line in lines]
    return '\n'.join(l for l in cleaned if l)


def remove_zero_width(text):
    """移除零宽字符、BOM 等不可见字符"""
    for ch in ('\u200b', '\u200c', '\u200d', '\ufeff', '\u2060'):
        text = text.replace(ch, '')
    return text


def sanitize(text):
    """综合文本清理"""
    return clean_whitespace(remove_zero_width(text))


def split_sentences(text):
    """将英文文本按句子切分。

    使用状态机跟踪引号开闭，确保：
    - 引号内部的 . ! ? 不会触发切分
    - 闭合引号后跟大写字母时才切分（新句开始）
    - 非引号环境下的 . ! ? 后跟大写字母/开引号时切分

    支持直引号 (") 和智能引号 (\u201c \u201d)。
    用于段落跟读和语篇跟读的逐句录音。
    """
    normalized = re.sub(r'\s+', ' ', text.strip())
    if not normalized:
        return []

    sentences = []
    current = ""
    in_quote = False          # 是否在引号内部
    quote_char = None         # 当前引号类型: '"' 或 '\u201c'

    i = 0
    n = len(normalized)
    while i < n:
        ch = normalized[i]

        # ---- 智能左引号 \u201c ----
        if ch == '\u201c':
            in_quote = True
            quote_char = '\u201c'
            current += ch
            i += 1
            continue

        # ---- 智能右引号 \u201d ----
        if ch == '\u201d':
            in_quote = False
            quote_char = None
            current += ch
            # 仅当引号内以 . ! ? 结尾时，才检查是否需要切分
            # （逗号结尾如 “Hello,” she said. 不切分）
            if len(current) >= 2 and current[-2] in '.!?':
                j = i + 1
                while j < n and normalized[j] == ' ':
                    j += 1
                if j < n and normalized[j].isupper():
                    sentences.append(current.strip())
                    current = ""
                    i = j
                    continue
            i += 1
            continue

        # ---- 直引号 " → 切换状态 ----
        if ch == '"':
            if not in_quote:
                # 开引号
                in_quote = True
                quote_char = '"'
                current += ch
                i += 1
                continue
            else:
                # 闭合引号
                in_quote = False
                quote_char = None
                current += ch
                # 仅当引号内以 . ! ? 结尾时，才检查是否需要切分
                # （逗号结尾如 "Hello," she said. 不切分）
                if len(current) >= 2 and current[-2] in '.!?':
                    j = i + 1
                    while j < n and normalized[j] == ' ':
                        j += 1
                    if j < n and normalized[j].isupper():
                        sentences.append(current.strip())
                        current = ""
                        i = j
                        continue
                i += 1
                continue

        # ---- 句末标点 . ! ? ----
        if ch in '.!?':
            current += ch
            if not in_quote:
                # 向前吞掉闭合括号/方括号（不吞引号，引号单独处理）
                j = i + 1
                while j < n and normalized[j] in ')]':
                    current += normalized[j]
                    j += 1
                # 跳过空白
                k = j
                while k < n and normalized[k] == ' ':
                    k += 1
                # 后面是大写字母或开引号 → 切分
                if k >= n:
                    # 文本结束，剩余部分由末尾兜底
                    i = k
                    continue
                if normalized[k].isupper() or normalized[k] in '\u201c"':
                    sentences.append(current.strip())
                    current = ""
                    i = k
                    continue
                # 没有切分 → 继续从 j 开始
                i = j
                continue
            # 引号内部：不切分，继续
            i += 1
            continue

        current += ch
        i += 1

    if current.strip():
        sentences.append(current.strip())

    if not sentences:
        sentences.append(normalized)
    return sentences


# ============================================================================
# 解析器基类
# ============================================================================

class BaseParser:
    """文档解析器基类，子类需实现 parse() 方法。"""

    DOC_TYPE = "未知"
    # 子类可设为 True 以跳过 load_paragraphs（如 Excel 解析器）
    _SKIP_LOAD_PARAGRAPHS = False

    def __init__(self, filepath):
        self.filepath = filepath
        self.filename = os.path.basename(filepath)
        if self._SKIP_LOAD_PARAGRAPHS:
            self.paras = []
        else:
            self.paras = load_paragraphs(filepath)

    def parse(self):
        """子类实现：返回解析结果字典"""
        raise NotImplementedError

    def _result(self, items):
        """构造标准输出结构"""
        return {
            "source_file": self.filename,
            "doc_type": self.DOC_TYPE,
            "item_count": len(items),
            "items": items,
        }


# ============================================================================
# 1. 信息获取解析器
# ============================================================================

class InfoAcquisitionParser(BaseParser):
    """
    解析「信息获取」文档。
    提取「第一节 听选信息」和「第二节 回答问题」中的题目及每段录音稿。

    题目规则：
      - 两节共用文档中的连续题号，题号不会在第二节重新开始
      - 每道题单独生成一条结果，文本去掉题号
      - 按题目出现顺序使用男声、女声交替朗读（第一题为男声）
      - filename_stem 使用"问题x"，供 TTS 阶段生成问题1.mp3等文件

    文档结构示例:
        第一节 听选信息
          听第一段对话，回答第1—2两个问题。
          1. Question?
          (Options)
          录音稿：
          W: ...
          M: ...
          听第二段对话，回答第3—4两个问题。
          ...
        第二节 回答问题
          听下面一段独白，录音播放两遍。
          7. Question?
          ...
          录音稿：
          (W) ...
        参考答案（独立一行）                ← 到此结束
        （新格式中每题后可能紧跟「参考答案：xxx」行内答案，
          该行不结束采集，会被跳过）
        （「回答问题」题目也可能漏编号，按上一个题号自动顺延）

    支持同一文件内包含多篇试题（多次出现「第一节 听选信息」）。
    """

    DOC_TYPE = "信息获取"

    # 第一节 听选信息
    RE_SECTION_START = re.compile(r'第一节\s*听选信息')
    # 第二节 回答问题
    RE_SECTION2_START = re.compile(r'第二节\s*回答问题')
    # 参考答案（仅独立一行）→ 结束所有收集。
    # 注意：新格式中每题后面紧跟「参考答案：xxx」的行不是结束标记，
    # 只有文档末尾独立成行的「参考答案：」才是。
    RE_SECTION_END = re.compile(r'^参考答案\s*[：:]?\s*$')
    # 行内参考答案（每题附带答案）→ 直接跳过
    RE_INLINE_ANSWER = re.compile(r'^参考答案\s*[：:]')
    # 漏编号的英文题目（新格式「回答问题」中 7-9 题可能没有题号）
    RE_UNNUMBERED_QUESTION = re.compile(r'[?？]\s*$')
    # 题号区间（如「回答第 7-10 个问题」「回答第1—2两个问题」），
    # 用于给漏编号题目确定起始题号
    RE_ANSWER_RANGE = re.compile(r'第\s*(\d+)\s*[-—~～至到]\s*(\d+)')
    # 任何「第X节」标记（用于检测其他题型的章节边界）
    RE_ANY_SECTION = re.compile(r'第[一二三四五六七八九十]+节')
    # 新旧题型混排时，兼容没有使用「第X节」而改用中文序号、Section
    # 或独立题型标题的边界；按结构识别，不把某个 Section 名称写死。
    RE_OTHER_MAJOR_HEADING = re.compile(
        r'^(?:[一二三四五六七八九十百]+\s*[、.．)]|'
        r'Section\s+[A-Z](?:\s*[：:]|$)|'
        r'听后选择(?:题型?|[（(【\s:：]|$))',
        re.I,
    )
    # 录音稿：（可能后面紧跟内容，也可能单独一行）
    RE_SCRIPT = re.compile(r'录音稿\s*[：:]\s*(.*)')
    # 听第X段对话 → 上一段录音稿结束
    RE_DIALOG = re.compile(r'听第.+段对话')
    # 题目编号：1. / 1． / 1、 / 1) / 1）
    RE_QUESTION = re.compile(r'^(\d+)\s*[.．、）)]\s*(.+)')

    def parse(self):
        items = []
        in_section = False       # 是否在任一节内
        current_category = ""    # 当前节的 category
        collecting = False       # 是否正在收集录音稿
        current_lines = []       # 当前录音稿的文本行
        question_order = 0       # 当前试题内的题目顺序；第二节不重置
        last_qnum = 0            # 上一个题目编号（用于漏编号题目顺延）
        next_qnum = None         # 下一个待分配的题号（由说明行的题号区间设定）
        # 每节独立编号
        idx_by_cat = {}          # {category: count}

        def flush():
            nonlocal collecting, current_lines
            if collecting and current_lines:
                cat = current_category
                idx_by_cat[cat] = idx_by_cat.get(cat, 0) + 1
                items.append({
                    "category": cat,
                    "index": idx_by_cat[cat],
                    "text": sanitize('\n'.join(current_lines)),
                })
            collecting = False
            current_lines = []

        for _, text, _ in self.paras:
            # ---- 第一节 听选信息 开始 ----
            if self.RE_SECTION_START.search(text):
                flush()
                in_section = True
                current_category = "听选信息录音稿"
                question_order = 0
                last_qnum = 0
                next_qnum = None
                continue

            # ---- 第二节 回答问题 开始 ----
            if self.RE_SECTION2_START.search(text):
                flush()
                in_section = True
                current_category = "回答问题录音稿"
                continue

            # ---- 其他「第X节」→ 遇到其他题型的章节，停止采集 ----
            if in_section and self.RE_ANY_SECTION.search(text):
                # 不是自己的章节标记，停止采集
                if not self.RE_SECTION_START.search(text) and not self.RE_SECTION2_START.search(text):
                    flush()
                    in_section = False
                    current_category = ""
                    continue

            if in_section and self.RE_OTHER_MAJOR_HEADING.match(text):
                flush()
                in_section = False
                current_category = ""
                continue

            # ---- 参考答案（独立一行）→ 结束所有收集 ----
            if in_section and self.RE_SECTION_END.match(text):
                flush()
                in_section = False
                current_category = ""
                continue

            if not in_section:
                continue

            # ---- 行内参考答案（每题附带答案）→ 跳过，不参与任何收集 ----
            if self.RE_INLINE_ANSWER.match(text):
                continue

            # ---- 题号区间说明行（如「回答第 7-10 个问题」）→ 设定起始题号 ----
            m_range = self.RE_ANSWER_RANGE.search(text)
            if m_range:
                next_qnum = int(m_range.group(1))

            # ---- 题目：去题号，按出现顺序男/女交替 ----
            # 只在非录音稿状态识别，避免把录音稿中偶然以数字开头的句子误判为题目。
            question_number = None
            question_text = ""
            if not collecting:
                m_q = self.RE_QUESTION.match(text)
                if m_q:
                    question_number = int(m_q.group(1))
                    question_text = sanitize(m_q.group(2))
                    # 有编号的题目推进待分配题号
                    next_qnum = max(next_qnum or 0, question_number + 1)
                elif self.RE_UNNUMBERED_QUESTION.search(text) and not is_chinese(text):
                    # 新格式中题目可能漏编号（如 U2 的 7-9 题）：
                    # 以英文问句（末尾 ?）兜底识别，题号按说明行区间顺延
                    question_number = (
                        next_qnum if next_qnum is not None else last_qnum + 1
                    )
                    if next_qnum is not None:
                        next_qnum += 1
                    question_text = sanitize(text)
            if question_number is not None and question_text:
                question_order += 1
                last_qnum = question_number
                speaker = "M" if question_order % 2 == 1 else "W"
                section_name = (
                    "听选信息" if current_category == "听选信息录音稿" else "回答问题"
                )
                items.append({
                    "category": f"{section_name}题目",
                    "number": question_number,
                    "filename_stem": f"问题{question_number}",
                    "voice": "male" if speaker == "M" else "female",
                    "text": f"{speaker}: {question_text}",
                })
                continue

            # ---- 录音稿标记 ----
            m = self.RE_SCRIPT.match(text)
            if m:
                flush()  # 先保存上一段录音稿
                collecting = True
                remainder = m.group(1).strip()
                if remainder:
                    current_lines.append(remainder)
                continue

            # ---- 新对话开始 → 当前录音稿结束 ----
            if collecting and self.RE_DIALOG.search(text):
                flush()
                continue

            # ---- 收集录音稿行 ----
            if collecting:
                current_lines.append(text)

        # 文档末尾兜底
        flush()

        return self._result(items)


# ============================================================================
# 2. 听后选择解析器
# ============================================================================

class ListeningSelectionParser(BaseParser):
    """只提取「听后选择」题型中的录音原文对话。

    这类试卷同时包含题干、选项和计算机提示，但配音任务只需要
    ``【录音原文】`` 后面的对话。W/w 使用默认女声，M/m 使用默认男声；
    标记本身保留在结构化文本中，后续合成阶段会负责切换音色并移除标记。
    """

    DOC_TYPE = "听后选择"

    # 题型标题可能带中文序号、题量说明或括号；只接受行首标题形态，
    # 避免正文里偶然提到“听后选择”时误启动解析状态机。
    RE_SECTION_START = re.compile(
        r'^(?:[一二三四五六七八九十百]+\s*[、.．)]\s*)?'
        r'听后选择(?:题型?|[（(【\s:：]|$)',
        re.I | re.M,
    )
    # 混合试卷中下一大节出现时，结束当前「听后选择」范围；同时兼容
    # 旧题型的「第X节」和新版 Section 标题。
    RE_MAJOR_SECTION = re.compile(
        r'^(?:[一二三四五六七八九十百]+\s*[、.．)]|'
        r'第[一二三四五六七八九十百]+节|'
        r'Section\s+[A-Z](?:\s*[：:]|$))',
        re.I,
    )
    # 支持【录音原文】、[录音原文]、（录音原文）以及普通冒号写法。
    RE_SCRIPT = re.compile(
        r'^[【\[（(]?\s*录音原文\s*[】\]）)]?\s*[：:]?\s*(.*)$'
    )
    RE_PROMPT = re.compile(
        r'计算机语音提示|语音提示|录音播放|现在，你有|听下面|请听录音|开始录音|停止录音'
    )
    RE_NUMBERED_QUESTION = re.compile(r'^\d+\s*[.．、）)]\s*[^A-Za-z]?')
    RE_OPTION = re.compile(r'^[A-CＡ-Ｃ]\s*[.．、）)]\s*')
    RE_ANSWER = re.compile(r'^(?:参考答案|答案|解析)\s*[：:]?')
    RE_SPEAKER_LINE = re.compile(r'^\s*([WwMm])\s*[:：]\s*(.*)$')

    @classmethod
    def _split_speaker_segments(cls, text):
        """把一段录音原文拆成按说话人轮次的独立音频文本。

        同一说话人连续出现时合并为一个段落；说话人切换时结束当前段落。
        保留 ``W:``/``M:`` 标记，供后续音色识别使用，但每个返回值都已经
        是一个最终音频条目，不能再把整道题重新拼回一个文件。
        """
        clean = sanitize(text)
        lines = clean.splitlines()
        labelled = [
            cls.RE_SPEAKER_LINE.match(line)
            for line in lines
            if str(line or '').strip()
        ]
        if not labelled or not any(labelled):
            return [(None, clean)] if clean else []

        segments = []
        current_role = None
        current_label = None
        current_lines = []

        def flush():
            nonlocal current_role, current_label, current_lines
            if not current_lines:
                return
            content = '\n'.join(line for line in current_lines if line).strip()
            if content:
                prefix = f"{current_label}: " if current_role else ""
                segments.append((current_role, f"{prefix}{content}"))
            current_role = None
            current_label = None
            current_lines = []

        for line in lines:
            value = str(line or '').strip()
            if not value:
                continue
            match = cls.RE_SPEAKER_LINE.match(value)
            if match:
                role = match.group(1).upper()
                if current_role is not None and role != current_role:
                    flush()
                if current_role is None:
                    current_role = role
                    current_label = match.group(1)
                content = match.group(2).strip()
                if content:
                    current_lines.append(content)
                continue

            # 标记后的换行是同一人的续行；标记前若有异常说明文字则单独
            # 保留，避免为了拆分而静默丢失原文。
            current_lines.append(value)

        flush()
        return segments

    def parse(self):
        items = []
        in_section = False
        collecting = False
        current_lines = []
        script_index = 0
        audio_index = 0

        def flush():
            nonlocal collecting, current_lines, script_index, audio_index
            text = sanitize('\n'.join(current_lines))
            if collecting and text:
                script_index += 1
                for speaker, segment_text in self._split_speaker_segments(text):
                    audio_index += 1
                    item = {
                        "category": "听后选择录音稿",
                        "index": audio_index,
                        "question_index": script_index,
                        "text": segment_text,
                    }
                    if speaker == "W":
                        item["speaker"] = "W"
                        item["voice"] = "female"
                    elif speaker == "M":
                        item["speaker"] = "M"
                        item["voice"] = "male"
                    items.append(item)
            collecting = False
            current_lines = []

        for _, text, _ in self.paras:
            value = str(text or '').strip()

            # 先处理题型标题，允许同一文档中出现多组听后选择。
            if self.RE_SECTION_START.search(value):
                flush()
                in_section = True
                continue

            if not in_section:
                continue

            # 下一大节属于其他题型时，停止采集，避免把后续录音原文混入。
            if self.RE_MAJOR_SECTION.match(value):
                flush()
                in_section = False
                continue

            script_match = self.RE_SCRIPT.match(value)
            if script_match:
                flush()
                collecting = True
                remainder = script_match.group(1).strip()
                if remainder:
                    current_lines.append(remainder)
                continue

            if not collecting:
                continue

            # 下一道题的提示、题干、选项或答案都不是录音内容。
            if (
                self.RE_PROMPT.search(value)
                or self.RE_NUMBERED_QUESTION.match(value)
                or self.RE_OPTION.match(value)
                or self.RE_ANSWER.match(value)
            ):
                flush()
                continue

            current_lines.append(value)

        flush()
        return self._result(items)


# ============================================================================
# 3. 听后应答解析器
# ============================================================================

class ListeningResponseParser(BaseParser):
    """解析「听后应答」中的计算机语音提示与待朗读句子。

    每个“（计算机语音提示）听下面 N 个句子。”提示开启一个采集块；
    采集块只包含该提示下方、直到“请朗读应答语”/倒计时/下一条提示
    之前的英文内容。内容按非空行输出为独立音频，全部使用默认女声。
    """

    DOC_TYPE = "听后应答"

    RE_SECTION_START = re.compile(r'听后应答')
    RE_PROMPT = re.compile(
        r'计算机语音提示.*?听下面\s*'
        r'(?P<count>[0-9０-９零〇一二两三四五六七八九十百]+)\s*个\s*句子',
        re.I,
    )
    RE_CONTROL = re.compile(
        r'计算机语音提示.*?(?:请朗读应答语|朗读应答语)|'
        r'计算机屏幕.*?(?:倒计时|进度条)|'
        r'(?:参考答案|答案|解析)\s*[：:]?',
        re.I,
    )
    RE_LEADING_MARK = re.compile(r'^[★☆*]\s*')
    RE_LEADING_NUMBER = re.compile(r'^\d+\s*[.．、）)]\s*')
    RE_GRADE = re.compile(
        r'(?<!\d)(?P<grade>[789七八九])\s*(?:年级\s*)?上',
        re.I,
    )

    _CHINESE_DIGITS = {
        '零': 0, '〇': 0, '一': 1, '二': 2, '两': 2, '三': 3,
        '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9,
    }
    _CHINESE_UNITS = {'十': 10, '百': 100}
    _GRADE_DIGITS = {'七': '7', '八': '8', '九': '9'}

    @classmethod
    def _parse_count(cls, value):
        """将阿拉伯数字或常见中文数字转换为句子数量。"""
        raw = str(value or '').strip().translate(str.maketrans(
            '０１２３４５６７８９', '0123456789'
        ))
        if raw.isdigit():
            return max(1, int(raw))
        if not raw:
            return None

        total = 0
        current = 0
        for char in raw:
            if char in cls._CHINESE_DIGITS:
                current = cls._CHINESE_DIGITS[char]
            elif char in cls._CHINESE_UNITS:
                unit = cls._CHINESE_UNITS[char]
                total += (current or 1) * unit
                current = 0
            else:
                return None
        result = total + current
        return max(1, result) if result > 0 else None

    @classmethod
    def _grade_prefix(cls, filename, paras):
        """从文件名或文档标题中提取 7上/8上/9上命名空间。"""
        haystack = ' '.join([
            str(filename or ''),
            *(str(text or '') for _, text, _ in paras[:8]),
        ])
        match = cls.RE_GRADE.search(haystack)
        if not match:
            return '应答'
        grade = match.group('grade')
        return f"{cls._GRADE_DIGITS.get(grade, grade)}上"

    @classmethod
    def _content_lines(cls, text):
        """清理提示块中的英文内容，并按换行保留音频边界。"""
        lines = []
        for raw_line in str(text or '').splitlines():
            value = sanitize(raw_line)
            if not value or cls.RE_CONTROL.search(value) or cls.RE_PROMPT.search(value):
                continue
            if is_chinese(value):
                continue
            value = cls.RE_LEADING_MARK.sub('', value)
            value = cls.RE_LEADING_NUMBER.sub('', value).strip()
            if value:
                lines.append(value)
        return lines

    def parse(self):
        items = []
        collecting = False
        expected_count = None
        current_lines = []
        response_index = 0
        grade_prefix = self._grade_prefix(self.filename, self.paras)

        def flush():
            nonlocal collecting, expected_count, current_lines, response_index
            for line in current_lines:
                response_index += 1
                items.append({
                    "category": "听后应答录音稿",
                    "index": response_index,
                    "number": response_index,
                    "filename_stem": f"{grade_prefix}-应答-{response_index}",
                    "voice": "female",
                    "text": line,
                })
            collecting = False
            expected_count = None
            current_lines = []

        for position, (_, text, _) in enumerate(self.paras):
            value = str(text or '').strip()
            prompt = self.RE_PROMPT.search(value)
            if prompt:
                flush()
                collecting = True
                expected_count = self._parse_count(prompt.group('count'))
                continue

            if not collecting:
                continue

            if self.RE_CONTROL.search(value):
                flush()
                continue

            current_lines.extend(self._content_lines(value))
            # 提示中的句数只作为边界提示，不能直接截断实际内容：题目可能
            # 把多行内容放在多个 Word 段落中，即使提示写着“1 个句子”，
            # 也必须遵循“多行多个音频”的规则。只有已经收集到提示数量，
            # 且下一个有内容的段落明确是控制提示/下一条提示时，才提前结束。
            if expected_count and len(current_lines) >= expected_count:
                next_value = ''
                for _, following_text, _ in self.paras[position + 1:]:
                    following_value = str(following_text or '').strip()
                    if following_value:
                        next_value = following_value
                        break
                if next_value and (
                    self.RE_PROMPT.search(next_value)
                    or self.RE_CONTROL.search(next_value)
                ):
                    flush()

        flush()
        return self._result(items)


# ============================================================================
# 4. 课文跟读解析器
# ============================================================================

class TextReadingParser(BaseParser):
    """
    解析「课文跟读」文档。
    提取三类内容：
      - 句子跟读：去掉序号前缀，按序号排序
      - 段落跟读：整段英文
      - 语篇跟读：按「语篇N」分组，每组可含多段

    Section A/B 新格式优先：显式 Conversation 对话按角色边界拆分，
    普通文章按标题/段落边界生成音频；旧版章节格式继续使用历史规则。

    音色/命名规则（仅 Understanding Idea / Reading for writing 章节）：
      讯飞声音参数由前端动态配置，解析器不固定。
      Understanding Idea (前缀 U)：
        - 句子跟读：男声，命名 U-句子1、U-句子2 …
        - 段落跟读：男声，逐句切分，命名 U-段落1、U-段落2 …
        - 语篇1：男声，逐句切分，命名 U-语篇1-1、U-语篇1-2 …
        - 语篇2：女声，逐句切分，命名 U-语篇2-1、U-语篇2-2 …
        - 语篇（仅一篇时）：男声，逐句切分，命名 U-语篇1-1、U-语篇1-2 …
      Reading for writing (前缀 R)：
        - 句子跟读：男声，命名 R-句子1、R-句子2 …
        - 段落跟读：男声，逐句切分，命名 R-段落1、R-段落2 …
        - 语篇跟读：女声，逐句切分，命名 R-语篇1-1、R-语篇1-2 …

    文档结构示例（两种标题层级均支持）：
        格式一（旧）：
        [Heading 1] 课文跟读-7上U1
        [Heading 2] Understanding Idea
          [Heading 3] 句子跟读
            1. English sentence
            中文：翻译
            2. English sentence
            ...
          [Heading 3] 段落跟读
            English paragraph
          [Heading 3] 语篇跟读
            语篇1
            English paragraph 1
            English paragraph 2
            语篇2
            ...
        [Heading 2] Reading for writing
          ...

        格式二（新）：
        [Normal] 课文跟读-7 上 U2
        [Heading 1] Understanding Idea
          [Heading 2] 句子跟读
            ...
          [Heading 2] 段落跟读
            ...
          [Heading 2] 语篇跟读
            English paragraph 1
            English paragraph 2
            （无显式「语篇N」标记，仅一篇时自动归为一组）
        [Heading 1] Reading for writing
          ...

    支持同一文件内包含多个单元/章节。
    """

    DOC_TYPE = "课文跟读"

    # 子章节标记（精确匹配）
    SUB_SENTENCE = "句子跟读"
    SUB_PARAGRAPH = "段落跟读"
    SUB_DISCOURSE = "语篇跟读"

    # 编号句：1. / 1、 / 1) / 1） 后跟内容
    RE_NUMBERED = re.compile(r'^(\d+)\s*[.、）)]\s*(.+)')
    # 语篇编号：语篇1 / 语篇 1 / 语篇1：
    RE_DISCOURSE_NUM = re.compile(r'^语篇\s*(\d+)\s*[：:]?\s*$')
    # 中文翻译行
    RE_CHINESE_PREFIX = re.compile(r'^中文\s*[：:]')
    # 已知的章节名（用于内容匹配，不依赖样式）
    KNOWN_SECTIONS = {
        'understanding idea', 'reading for writing',
        'developing ideas', 'developing idea',
    }

    @staticmethod
    def _section_prefix(section):
        """根据章节名返回命名前缀，未知章节返回空字符串。"""
        s = section.strip().lower()
        if 'understanding' in s:
            return "U"
        if 'reading' in s:
            return "R"
        return ""

    @classmethod
    def _discourse_voice(cls, section, discourse_number):
        """确定语篇跟读的音色。
        Understanding Idea: 语篇1→男声, 语篇2→女声
        Reading for writing: 全部女声
        """
        prefix = cls._section_prefix(section)
        if prefix == "U":
            return "male" if discourse_number == 1 else "female"
        if prefix == "R":
            return "female"
        return "female"

    # 子章节名称集合，用于排除 heading 样式误匹配
    _SUB_SECTION_NAMES = frozenset({SUB_SENTENCE, SUB_PARAGRAPH, SUB_DISCOURSE})

    # ------------------------------------------------------------------
    # Section A/B 新版课文跟读格式
    # ------------------------------------------------------------------
    # 新版样本中 Section A/B 和子题型有时只是 Normal 样式，不能依赖
    # Heading 层级判断；而且「段落/语篇跟读」的音频边界由 Conversation
    # 或 Word 段落决定，不能套用旧版的逐句切分规则。
    RE_SECTION_AB = re.compile(r'^Section\s+([A-Za-z])\s*[：:]?\s*$', re.I)
    RE_CONVERSATION = re.compile(r'^Conversation\s*(\d+)\s*[：:]?\s*$', re.I)
    RE_NEW_SUB_SECTION = re.compile(r'^(句子跟读|段落跟读|语篇跟读)\s*[：:]?\s*$')
    RE_ROLE_LABEL = re.compile(r'^([^:：\n]{1,60}?)\s*[:：]\s*(.*)$')
    RE_ARTICLE_SUBTITLE = re.compile(r'^//\s*(.*)$')
    RE_CONTENT_HEADING = re.compile(r'^(Reading\s+Plus)\s*[：:]?\s*$', re.I)

    def __init__(self, filepath):
        super().__init__(filepath)
        self._section_ab_profile_cache = None

    def parse(self):
        """按格式优先级解析课文跟读。

        新版不是由某一个标题字符串单独决定的：只有 Section 标记、跟读
        子题型和有效内容形成连续结构时，才切换到新版解析；否则继续走
        原有 Understanding Idea / Reading for writing 逻辑，避免旧文档中
        偶然出现的 ``Section A`` 被误判。
        """
        if self._is_section_ab_format():
            return self._parse_section_ab_format()
        return self._parse_legacy_format()

    def _is_section_ab_format(self):
        """判断是否为新版 Section A/B 课文跟读文档。

        Section A/B 只是候选信号。真正的判定要求至少有一个 Section
        标记，后面紧跟一个跟读子题型，并且该子题型下存在可朗读内容。
        这样旧文档正文、目录或备注中的孤立 ``Section A`` 不会触发新版
        状态机；如果新旧结构都完整存在，则按产品要求让新版优先。
        """
        return self._detect_section_ab_profile()["is_new"]

    @classmethod
    def _normalized_heading(cls, text):
        """统一标题末尾中英文冒号和空白，供结构检测使用。"""
        value = str(text or '').strip()
        return re.sub(r'[：:]\s*$', '', value).strip()

    @classmethod
    def _new_subsection_name(cls, text):
        """返回新版跟读子题型名；普通正文返回 None。"""
        value = cls._normalized_heading(text)
        return value if value in cls._SUB_SECTION_NAMES else None

    @classmethod
    def _is_reading_plus_heading(cls, text):
        """只做内容层面的候选匹配，是否为结构标题由 profile 决定。"""
        return bool(cls.RE_CONTENT_HEADING.match(str(text or '').strip()))

    def _detect_section_ab_profile(self):
        """从段落序列提取新版格式证据，不依赖文件名或年级名称。

        ``Section A`` 只有在同一 Section 范围内连接到跟读子题型，且子
        题型下有英文/角色等有效载荷时才算成立。这里还同时提取
        Reading Plus 的结构位置和对话的角色拆分模式，解析阶段复用这份
        profile，避免不同分支各自用启发式重新猜测。
        """
        if self._section_ab_profile_cache is not None:
            return self._section_ab_profile_cache

        entries = list(self.paras)
        section_positions = []
        legacy_section_positions = []
        subsection_positions = []
        reading_plus_candidates = []
        conversation_positions = []
        role_line_positions = []
        article_marker_positions = []

        for position, (_, text, _) in enumerate(entries):
            value = str(text or '').strip()
            if self.RE_SECTION_AB.match(value):
                section_positions.append(position)
            if self._normalized_heading(value).casefold() in self.KNOWN_SECTIONS:
                legacy_section_positions.append(position)
            subsection = self._new_subsection_name(value)
            if subsection:
                subsection_positions.append((position, subsection))
            if self._is_reading_plus_heading(value):
                reading_plus_candidates.append(position)
            if self.RE_CONVERSATION.match(value):
                conversation_positions.append(position)
            if self.RE_ARTICLE_SUBTITLE.match(value):
                article_marker_positions.append(position)
            if any(self._role_label(line) for line in value.splitlines()):
                role_line_positions.append(position)

        def next_section_after(position):
            return next(
                (
                    candidate
                    for candidate in sorted(section_positions + legacy_section_positions)
                    if candidate > position
                ),
                len(entries),
            )

        # Section 与子题型必须在同一个 Section 范围内，避免文档前言中的
        # Section A 和后面完全无关的跟读标题被拼成新版结构。
        linked_subsections = []
        for section_position in section_positions:
            section_end = next_section_after(section_position)
            linked_subsections.extend(
                (position, name)
                for position, name in subsection_positions
                if section_position < position < section_end
            )

        # 仅把跟读子题型之后、下一个结构标题之前的英文/角色内容算作
        # 有效载荷。标题、Conversation 编号、语篇编号本身不算内容。
        payload_subsections = []
        all_boundary_positions = sorted(
            section_positions
            + legacy_section_positions
            + [position for position, _ in subsection_positions]
            + reading_plus_candidates
        )
        for subsection_position, subsection_name in linked_subsections:
            next_boundary = next(
                (candidate for candidate in all_boundary_positions
                 if candidate > subsection_position),
                len(entries),
            )
            has_payload = False
            for position in range(subsection_position + 1, next_boundary):
                value = str(entries[position][1] or '').strip()
                if not value:
                    continue
                if self.RE_CONVERSATION.match(value):
                    continue
                if self.RE_DISCOURSE_NUM.match(value):
                    continue
                if self._is_reading_plus_heading(value):
                    continue
                if self._new_english_lines(value):
                    has_payload = True
                    break
            if has_payload:
                payload_subsections.append((subsection_position, subsection_name))

        # Reading Plus 只有在短距离内确实引出新的跟读子题型时才视为章节
        # 边界；普通正文中提到这个词不会改变前面的 Section 和命名空间。
        reading_plus_positions = set()
        for candidate in reading_plus_candidates:
            candidate_end = next_section_after(candidate)
            if any(
                candidate < subsection_position <= candidate + 8
                and subsection_position < candidate_end
                for subsection_position, _ in subsection_positions
            ):
                reading_plus_positions.add(candidate)

        # 对话模式同样从结构判断：显式 Conversation 块中出现至少两个
        # 有效角色时，说明文档给出了“按角色分块”的边界。没有显式
        # Conversation 的新格式对话，则在解析阶段按 Word 段落保留边界，
        # 由 flush_dialogue_buffer 进一步拆成逐行音频；普通非对话段落不受影响。
        conversation_blocks = []
        active_block = None
        current_subsection = None
        in_new_section = False

        def flush_conversation_block():
            nonlocal active_block
            if active_block is not None:
                conversation_blocks.append(active_block)
                active_block = None

        for position, (_, text, _) in enumerate(entries):
            value = str(text or '').strip()
            section_match = self.RE_SECTION_AB.match(value)
            if section_match:
                flush_conversation_block()
                current_subsection = None
                in_new_section = True
                continue
            if self._normalized_heading(value).casefold() in self.KNOWN_SECTIONS:
                flush_conversation_block()
                current_subsection = None
                in_new_section = False
                continue
            if not in_new_section:
                continue
            subsection = self._new_subsection_name(value)
            if subsection:
                flush_conversation_block()
                current_subsection = subsection
                continue
            if self.RE_CONVERSATION.match(value):
                flush_conversation_block()
                if current_subsection in (self.SUB_PARAGRAPH, self.SUB_DISCOURSE):
                    active_block = {
                        "position": position,
                        "roles": set(),
                        "role_lines": 0,
                    }
                continue
            if active_block is not None:
                for line in self._new_english_lines(value):
                    role = self._role_label(line)
                    if role:
                        active_block["roles"].add(
                            re.sub(r'\s+', ' ', role).casefold()
                        )
                        active_block["role_lines"] += 1
        flush_conversation_block()

        role_dialogue_blocks = [
            block for block in conversation_blocks
            if len(block["roles"]) >= 2 and block["role_lines"] >= 2
        ]
        role_audio_mode = "per_role" if role_dialogue_blocks else "aggregate"

        # 新版置信条件：Section 标记、同范围跟读子题型、子题型载荷三者
        # 缺一不可。没有使用单个字符串、文件名、标题或年级名做决策。
        is_new = bool(
            section_positions
            and linked_subsections
            and payload_subsections
        )
        self._section_ab_profile_cache = {
            "is_new": is_new,
            "section_positions": tuple(section_positions),
            "legacy_section_positions": tuple(legacy_section_positions),
            "subsection_positions": tuple(subsection_positions),
            "reading_plus_positions": frozenset(reading_plus_positions),
            "conversation_positions": tuple(conversation_positions),
            "role_line_positions": tuple(role_line_positions),
            "article_marker_positions": tuple(article_marker_positions),
            "role_audio_mode": role_audio_mode,
            "conversation_blocks": tuple(conversation_blocks),
        }
        return self._section_ab_profile_cache

    @staticmethod
    def _role_label_is_valid(label):
        """判断冒号前文本是否更像角色名，而不是普通正文。"""
        value = str(label or '').strip()
        if not value or len(value) > 48:
            return False
        if len(re.split(r'\s+', value)) > 4:
            return False
        if value[0].isdigit() or '://' in value or '/' in value or '\\' in value:
            return False
        if re.search(r'[.!?。！？；;，,]', value):
            return False
        return True

    @classmethod
    def _role_label(cls, text):
        """返回一行中的角色名；普通带冒号文本返回 None。"""
        match = cls.RE_ROLE_LABEL.match(str(text or '').strip())
        if not match or not cls._role_label_is_valid(match.group(1)):
            return None
        return match.group(1).strip()

    @classmethod
    def _contains_multiple_roles(cls, texts):
        """新版对话至少出现两个不同角色时，按对话整体保留换行。"""
        labels = set()
        for text in texts:
            for line in str(text or '').splitlines():
                label = cls._role_label(line)
                if label:
                    labels.add(re.sub(r'\s+', ' ', label).casefold())
        return len(labels) >= 2

    @staticmethod
    def _new_english_lines(text):
        """从新版内容中去掉中文翻译行，保留英文/角色行。"""
        lines = []
        for line in str(text or '').splitlines():
            line = line.strip()
            if not line:
                continue
            if TextReadingParser.RE_CHINESE_PREFIX.match(line) or is_chinese(line):
                continue
            lines.append(line)
        return lines

    @classmethod
    def _new_clean_text(cls, texts):
        """清理新版音频文本，同时保留角色行之间的换行。"""
        lines = []
        for text in texts:
            lines.extend(cls._new_english_lines(text))
        return sanitize('\n'.join(lines))

    @classmethod
    def _new_article_units(cls, text):
        """按新版文章中的 ``//`` 小标题标记拆分同一 Word 段落。"""
        units = []
        current_lines = []
        for line in cls._new_english_lines(text):
            if cls.RE_ARTICLE_SUBTITLE.match(line):
                if current_lines:
                    units.append('\n'.join(current_lines))
                    current_lines = []
                # 双斜杠本身是结构标记，不与下一行正文混为一个原始段落；
                # 后续 append_article_items 会把它和紧邻段落合并成一个音频。
                units.append(line)
                continue
            current_lines.append(line)
        if current_lines:
            units.append('\n'.join(current_lines))
        return units

    @classmethod
    def _article_heading(cls, text):
        """返回文章标题/小标题文本及是否显式使用 // 标记。"""
        value = sanitize(text)
        if not value:
            return '', False
        match = cls.RE_ARTICLE_SUBTITLE.match(value)
        if match:
            return match.group(1).strip(), True
        return value, False

    @classmethod
    def _looks_like_article_heading(cls, text):
        """识别新规则中的文章大标题/小标题。

        新样本的标题使用短文本、粗体或字号区分，但解析主流程只保留
        统一的段落文本；短文本且没有句末标点是对未带 // 标记样本的
        稳定兜底规则，显式 // 始终优先。
        """
        value, explicit = cls._article_heading(text)
        if not value:
            return False
        if explicit:
            return True
        if cls._role_label(value):
            return False
        if len(value) > 80 or len(re.split(r'\s+', value)) > 12:
            return False
        if re.match(r'^\d+\s*[.、）)]', value):
            return False
        if re.search(r'[.!?。！？]$', value):
            return False
        return True

    @staticmethod
    def _section_ab_code(section):
        match = re.match(r'^Section\s+([A-Za-z])$', str(section or '').strip(), re.I)
        return f"S{match.group(1).upper()}" if match else "S"

    @classmethod
    def _role_segments(cls, text):
        """按角色行拆分一条对话，返回 ``[(角色名, 文本), ...]``。"""
        segments = []
        current_role = None
        current_lines = []

        def flush():
            nonlocal current_lines
            clean = sanitize('\n'.join(current_lines))
            if clean:
                segments.append((current_role, clean))
            current_lines = []

        for line in str(text or '').splitlines():
            value = line.strip()
            if not value:
                continue
            role = cls._role_label(value)
            if role:
                flush()
                current_role = role
            current_lines.append(value)
        flush()
        return segments or [(None, sanitize(text))]

    def _parse_section_ab_format(self):
        """解析讯飞新版 Section A/B 课文跟读格式。

        新版规则：
          - 句子跟读按编号输出，默认女声；
          - 没有显式 Conversation 的多角色段落按 Word 段落/角色行输出，
            角色名通过结构元数据提供给用户配置；显式 Conversation 块中
            每个角色单独输出，拆分模式由文档结构 profile 决定；
          - 语篇跟读的对话遵循同样的角色拆分规则；文章按段落输出，
            大标题单独一个音频，// 小标题与其紧邻的一个段落合并。
        """
        format_profile = self._detect_section_ab_profile()
        split_role_audio = format_profile["role_audio_mode"] == "per_role"
        reading_plus_positions = format_profile["reading_plus_positions"]
        items = []
        current_section = ''
        current_sub = None
        current_audio_prefix = 'S'
        in_new_section = False

        sentence_buf = []
        paragraph_units = []
        paragraph_blocks = []
        paragraph_conversation_mode = False
        paragraph_current_number = None
        paragraph_current_lines = []

        discourse_units = []
        discourse_blocks = []
        discourse_conversation_mode = False
        discourse_current_number = None
        discourse_current_lines = []
        new_sequence_by_category = {}

        def reset_paragraph_state():
            nonlocal paragraph_units, paragraph_blocks
            nonlocal paragraph_conversation_mode, paragraph_current_number
            nonlocal paragraph_current_lines
            paragraph_units = []
            paragraph_blocks = []
            paragraph_conversation_mode = False
            paragraph_current_number = None
            paragraph_current_lines = []

        def reset_discourse_state():
            nonlocal discourse_units, discourse_blocks
            nonlocal discourse_conversation_mode, discourse_current_number
            nonlocal discourse_current_lines
            discourse_units = []
            discourse_blocks = []
            discourse_conversation_mode = False
            discourse_current_number = None
            discourse_current_lines = []

        def reset_section_sequences():
            new_sequence_by_category.clear()

        def flush_sentences():
            nonlocal sentence_buf
            if not sentence_buf:
                return
            for number, text in sorted(sentence_buf, key=lambda value: value[0]):
                items.append({
                    "category": self.SUB_SENTENCE,
                    "section": current_section,
                    "number": number,
                    "filename_stem": f"{current_audio_prefix}句子{number}",
                    "voice": "female",
                    "text": text,
                })
            sentence_buf = []

        def append_new_block_item(category, text, conversation_number=None, role=None):
            clean = sanitize(text)
            if not clean:
                return
            sequence_key = f"{current_audio_prefix}:{category}"
            new_sequence_by_category[sequence_key] = new_sequence_by_category.get(sequence_key, 0) + 1
            sequence = new_sequence_by_category[sequence_key]
            if (
                category in (self.SUB_PARAGRAPH, self.SUB_DISCOURSE)
                and conversation_number is not None
            ):
                # 新题型要求显式 Conversation 的段落/语篇跟读保留对话编号：
                # SA-段-Cx-y / SA-语-Cx-y，x 为 Conversation 编号，y 为
                # 当前题型内的音频生成序号。
                mode_prefix = "段" if category == self.SUB_PARAGRAPH else "语"
                filename_stem = (
                    f"{current_audio_prefix}-{mode_prefix}-C"
                    f"{conversation_number}-{sequence}"
                )
            else:
                filename_stem = f"{current_audio_prefix}{category[:2]}{sequence}"
            item = {
                "category": category,
                "section": current_section,
                "number": sequence,
                "filename_stem": filename_stem,
                "text": clean,
            }
            # 对话可能有多个角色，不能给整条结果写死男女声；未知角色
            # 在合成阶段按默认女声处理，已选择的角色由 role_voices 覆盖。
            if role:
                item["role"] = role
            elif not self._contains_multiple_roles(clean.splitlines()):
                item["voice"] = "female"
            if conversation_number is not None:
                item["conversation_number"] = conversation_number
            items.append(item)

        def flush_dialogue_buffer(category, units, blocks, conversation_mode):
            """输出新版段落/语篇对话。"""
            if conversation_mode:
                groups = [
                    (conversation_number, text, None)
                    for conversation_number, text in blocks
                ]
            else:
                # 新版“段落跟读”的对话通常没有 Conversation 标记，而是
                # 每个 Word 段落一行角色台词。检测到至少两个角色后，必须
                # 保留这些段落边界，否则整个对话会被错误合成一条音频。
                # 非对话文本仍沿用原来的“一段 Word 段落一条音频”规则。
                groups = []
                is_dialogue = self._contains_multiple_roles(units)
                for unit in units:
                    role_segments = self._role_segments(unit)
                    if is_dialogue and len(role_segments) > 1:
                        groups.extend(
                            (None, role_text, role)
                            for role, role_text in role_segments
                        )
                        continue
                    role_hint = (
                        role_segments[0][0]
                        if is_dialogue and len(role_segments) == 1
                        else None
                    )
                    groups.append((
                        None,
                        self._new_clean_text([unit]),
                        role_hint,
                    ))
            for conversation_number, text, role_hint in groups:
                if isinstance(text, list):
                    text = self._new_clean_text(text)
                if (
                    split_role_audio
                    and conversation_mode
                    and self._contains_multiple_roles([text])
                ):
                    for role, role_text in self._role_segments(text):
                        append_new_block_item(
                            category,
                            role_text,
                            conversation_number,
                            role=role,
                        )
                else:
                    append_new_block_item(
                        category,
                        text,
                        conversation_number,
                        role=role_hint,
                    )

        def flush_paragraph():
            nonlocal paragraph_current_lines, paragraph_current_number
            if paragraph_conversation_mode and paragraph_current_lines:
                paragraph_blocks.append((
                    paragraph_current_number,
                    self._new_clean_text(paragraph_current_lines),
                ))
                paragraph_current_lines = []
                paragraph_current_number = None
            if paragraph_conversation_mode:
                flush_dialogue_buffer(
                    self.SUB_PARAGRAPH,
                    paragraph_units,
                    paragraph_blocks,
                    True,
                )
            elif paragraph_units:
                flush_dialogue_buffer(
                    self.SUB_PARAGRAPH,
                    paragraph_units,
                    [],
                    False,
                )
            reset_paragraph_state()

        def append_article_items(units):
            """按文章标题/小标题规则输出语篇音频。"""
            cleaned = []
            for unit in units:
                value = self._new_clean_text([unit])
                if value:
                    cleaned.append(value)
            if not cleaned:
                return

            # 第一个短标题是文章大标题，单独一个音频；显式 // 作为
            # 小标题时不触发大标题判断，后续按小标题+下一段合并。
            index = 0
            first_value, first_explicit = self._article_heading(cleaned[0])
            if len(cleaned) > 1 and self._looks_like_article_heading(cleaned[0]) and not first_explicit:
                append_new_block_item(self.SUB_DISCOURSE, first_value)
                index = 1

            while index < len(cleaned):
                value, explicit = self._article_heading(cleaned[index])
                if explicit or self._looks_like_article_heading(value):
                    if index + 1 < len(cleaned):
                        next_value = cleaned[index + 1]
                        append_new_block_item(
                            self.SUB_DISCOURSE,
                            f"{value}\n{next_value}",
                        )
                        index += 2
                    else:
                        append_new_block_item(self.SUB_DISCOURSE, value)
                        index += 1
                    continue
                append_new_block_item(self.SUB_DISCOURSE, cleaned[index])
                index += 1

        def flush_discourse():
            nonlocal discourse_current_lines, discourse_current_number
            if discourse_conversation_mode and discourse_current_lines:
                discourse_blocks.append((
                    discourse_current_number,
                    self._new_clean_text(discourse_current_lines),
                ))
                discourse_current_lines = []
                discourse_current_number = None
            if discourse_conversation_mode:
                flush_dialogue_buffer(
                    self.SUB_DISCOURSE,
                    discourse_units,
                    discourse_blocks,
                    True,
                )
            elif discourse_units:
                # 兼容旧式「语篇1」标记：标记本身不朗读，组内仍按新版
                # “一段一个音频/标题规则”处理。
                groups = []
                current = []
                for unit in discourse_units:
                    if self.RE_DISCOURSE_NUM.match(unit):
                        if current:
                            groups.append(current)
                            current = []
                        continue
                    current.append(unit)
                if current:
                    groups.append(current)
                if not groups:
                    groups = [discourse_units]
                for group in groups:
                    append_article_items(group)
            reset_discourse_state()

        def flush_all_new():
            flush_sentences()
            flush_paragraph()
            flush_discourse()

        def append_new_content(target, text):
            lines = self._new_english_lines(text)
            if not lines:
                return
            if target == self.SUB_PARAGRAPH:
                if paragraph_conversation_mode:
                    paragraph_current_lines.extend(lines)
                else:
                    paragraph_units.append('\n'.join(lines))
            elif target == self.SUB_DISCOURSE:
                if discourse_conversation_mode:
                    discourse_current_lines.extend(lines)
                else:
                    discourse_units.extend(self._new_article_units(text))

        for position, (_, text, _) in enumerate(self.paras):
            section_match = self.RE_SECTION_AB.match(text)
            if section_match:
                flush_all_new()
                current_section = f"Section {section_match.group(1).upper()}"
                current_audio_prefix = self._section_ab_code(current_section)
                reset_section_sequences()
                current_sub = None
                in_new_section = True
                continue

            # 混合文档中如果在新版 Section 之间出现完整的旧章节，先把
            # 它从新版状态机隔离出来，避免旧章节的「句子跟读」被错误
            # 命名成 S句子N。新版分支只处理已经确认属于新版的 Section
            # 范围；完整旧格式文档仍由 legacy 入口处理。
            if self._normalized_heading(text).casefold() in self.KNOWN_SECTIONS:
                flush_all_new()
                current_section = ''
                current_audio_prefix = 'S'
                reset_section_sequences()
                current_sub = None
                in_new_section = False
                continue

            if not in_new_section:
                continue

            # 样本中的 Reading Plus 是语篇跟读下一个内容组的标题，
            # 不属于上一篇文章的音频文本；下一个「语篇跟读」会再次建立
            # 音频边界。样式可能是 Normal，因此用内容标记兜底。
            if position in reading_plus_positions:
                flush_all_new()
                current_section = "Reading Plus"
                current_audio_prefix = "RP"
                reset_section_sequences()
                current_sub = None
                continue

            sub_match = self.RE_NEW_SUB_SECTION.match(text)
            if sub_match:
                flush_all_new()
                current_sub = sub_match.group(1)
                continue

            if current_sub == self.SUB_SENTENCE:
                self._handle_sentence(text, sentence_buf)
                continue

            if current_sub == self.SUB_PARAGRAPH:
                conversation_match = self.RE_CONVERSATION.match(text)
                if conversation_match:
                    if paragraph_conversation_mode and paragraph_current_lines:
                        paragraph_blocks.append((
                            paragraph_current_number,
                            self._new_clean_text(paragraph_current_lines),
                        ))
                    paragraph_conversation_mode = True
                    paragraph_current_number = int(conversation_match.group(1))
                    paragraph_current_lines = []
                    continue
                append_new_content(self.SUB_PARAGRAPH, text)
                continue

            if current_sub == self.SUB_DISCOURSE:
                conversation_match = self.RE_CONVERSATION.match(text)
                if conversation_match:
                    if discourse_conversation_mode and discourse_current_lines:
                        discourse_blocks.append((
                            discourse_current_number,
                            self._new_clean_text(discourse_current_lines),
                        ))
                    discourse_conversation_mode = True
                    discourse_current_number = int(conversation_match.group(1))
                    discourse_current_lines = []
                    continue
                append_new_content(self.SUB_DISCOURSE, text)

        flush_all_new()
        return self._result(items)

    def _parse_legacy_format(self):
        items = []
        current_sub = None        # 当前子章节类型
        current_section = ""      # 当前章节名

        # 句子跟读临时存储: [(序号, 英文), ...]
        sentence_buf = []
        # 段落跟读临时存储: [行, ...]
        paragraph_buf = []
        # 语篇跟读临时存储: {序号: [行, ...]}
        discourse_buf = {}

        def flush_sentences():
            nonlocal sentence_buf
            if sentence_buf:
                prefix = self._section_prefix(current_section)
                sentence_buf.sort(key=lambda x: x[0])
                for num, text in sentence_buf:
                    item = {
                        "category": "句子跟读",
                        "section": current_section,
                        "number": num,
                        "text": text,
                    }
                    if prefix:
                        item["voice"] = "male"
                        item["filename_stem"] = f"{prefix}-句子{num}"
                    items.append(item)
                sentence_buf = []

        def flush_paragraph():
            nonlocal paragraph_buf
            if paragraph_buf:
                prefix = self._section_prefix(current_section)
                full_text = sanitize('\n'.join(paragraph_buf))
                # 逐句切分，每句一个音频文件
                sentences = split_sentences(full_text)
                if not sentences:
                    sentences = [full_text]
                for sent_idx, sent_text in enumerate(sentences, 1):
                    item = {
                        "category": "段落跟读",
                        "section": current_section,
                        "sentence_number": sent_idx,
                        "text": sent_text,
                    }
                    if prefix:
                        item["voice"] = "male"
                        item["filename_stem"] = f"{prefix}-段落{sent_idx}"
                    items.append(item)
                paragraph_buf = []

        def flush_discourse():
            nonlocal discourse_buf
            if discourse_buf:
                prefix = self._section_prefix(current_section)
                discourse_count = len(discourse_buf)
                for num in sorted(discourse_buf.keys()):
                    lines = discourse_buf[num]
                    if not lines:
                        continue
                    full_text = sanitize('\n'.join(lines))
                    # 确定音色和显示编号
                    if prefix == "U":
                        if discourse_count == 1:
                            voice = "male"
                            display_num = 1  # 仅一篇时也用「语篇1」
                        else:
                            voice = self._discourse_voice(current_section, num)
                            display_num = num
                    elif prefix == "R":
                        voice = self._discourse_voice(current_section, num)
                        display_num = 1  # R 章节仅一篇，用「语篇1」
                    else:
                        voice = "female"
                        display_num = max(num, 1)
                    # 逐句切分，每句一个音频文件
                    sentences = split_sentences(full_text)
                    if not sentences:
                        sentences = [full_text]
                    for sent_idx, sent_text in enumerate(sentences, 1):
                        item = {
                            "category": "语篇跟读",
                            "section": current_section,
                            "discourse_number": num,
                            "sentence_number": sent_idx,
                            "text": sent_text,
                        }
                        if prefix:
                            item["voice"] = voice
                            item["filename_stem"] = (
                                f"{prefix}-语篇{display_num}-{sent_idx}"
                            )
                        items.append(item)
                discourse_buf = {}

        def flush_all():
            flush_sentences()
            flush_paragraph()
            flush_discourse()

        for _, text, style in self.paras:
            # ---- 子章节标记（优先于 Heading 检查，兼容新格式中
            #      子章节使用 Heading 2 样式的情况）----
            if text == self.SUB_SENTENCE:
                flush_all()
                current_sub = self.SUB_SENTENCE
                continue
            if text == self.SUB_PARAGRAPH:
                flush_all()
                current_sub = self.SUB_PARAGRAPH
                continue
            if text == self.SUB_DISCOURSE:
                flush_all()
                current_sub = self.SUB_DISCOURSE
                continue

            # ---- Heading：新章节 ----
            if self._is_section_heading(style, text):
                flush_all()
                current_section = text
                current_sub = None
                continue

            # ---- 根据子章节处理内容 ----
            if current_sub == self.SUB_SENTENCE:
                self._handle_sentence(text, sentence_buf)

            elif current_sub == self.SUB_PARAGRAPH:
                # 跳过中文翻译行
                if not self.RE_CHINESE_PREFIX.match(text):
                    paragraph_buf.append(text)

            elif current_sub == self.SUB_DISCOURSE:
                self._handle_discourse(text, discourse_buf)

        # 文档末尾兜底
        flush_all()
        return self._result(items)

    def _is_section_heading(self, style, text):
        """判断是否为章节标题（Understanding Idea / Reading for writing 等）。

        兼容两种文档格式：
        - 旧格式：章节为 Heading 2，子章节为 Heading 3
        - 新格式：章节为 Heading 1，子章节为 Heading 2

        策略：
        1. 已知章节名直接匹配（不依赖样式）
        2. Heading 1/2 样式且不是子章节名时，也视为章节标题
        """
        # 已知章节名直接匹配
        if text.strip().lower() in self.KNOWN_SECTIONS:
            return True
        # Heading 1 或 Heading 2 样式（排除子章节名，避免误匹配）
        if style:
            sl = style.lower()
            if ('heading 1' in sl or 'heading 2' in sl):
                if text.strip() not in self._SUB_SECTION_NAMES:
                    return True
        return False

    def _handle_sentence(self, text, buf):
        """处理句子跟读的一个段落

        支持两种格式：
        - 带编号：1. English sentence / 1、English sentence
        - 无编号：English sentence（如 Understanding Idea 中的格式，
          每个英文句子后跟一行中文翻译）
        """
        # 跳过中文翻译行
        if self.RE_CHINESE_PREFIX.match(text):
            return
        # 取第一行（段落内可能内嵌中文翻译换行）
        first_line = text.split('\n')[0].strip()
        m = self.RE_NUMBERED.match(first_line)
        if m:
            num = int(m.group(1))
            eng = m.group(2).strip()
            # 跳过纯中文行
            if eng and not is_chinese(eng):
                buf.append((num, eng))
        else:
            # 无编号格式：英文句子后跟中文翻译行
            # 跳过纯中文行，保留英文句子
            if not is_chinese(text):
                eng = sanitize(text)
                if eng:
                    # 自动编号：基于当前缓冲区大小
                    num = len(buf) + 1
                    buf.append((num, eng))

    def _handle_discourse(self, text, buf):
        """处理语篇跟读的一个段落"""
        m = self.RE_DISCOURSE_NUM.match(text)
        if m:
            num = int(m.group(1))
            if num not in buf:
                buf[num] = []
        else:
            # 跳过中文翻译行，将英文内容追加到当前语篇
            if not self.RE_CHINESE_PREFIX.match(text) and buf:
                last_key = max(buf.keys())
                buf[last_key].append(text)
            elif not self.RE_CHINESE_PREFIX.match(text):
                # 没有显式「语篇N」标记时，归入默认组 0
                if 0 not in buf:
                    buf[0] = []
                buf[0].append(text)


# ============================================================================
# 5. 信息转述及询问解析器
# ============================================================================

class InfoRetellingParser(BaseParser):
    """
    解析「信息转述及询问」文档。
    提取「第一节 信息转述」的录音稿。

    文档结构示例:
        第一节 信息转述
          ... (题目描述)
          录音稿：(W) I'm Emma...     ← 录音稿可能与标记在同一行
          参考答案：...               ← 到此结束
        第二节 询问信息               ← 或到这里结束

    支持同一文件内包含多篇试题。
    """

    DOC_TYPE = "信息转述及询问"

    RE_SECTION_START = re.compile(r'第一节\s*信息转述')
    RE_SECTION_END = re.compile(r'第二节|参考答案')
    # 任何「第X节」标记（用于检测其他题型的章节边界）
    RE_ANY_SECTION = re.compile(r'第[一二三四五六七八九十]+节')
    RE_SCRIPT = re.compile(r'录音稿\s*[：:]\s*(.*)')

    def parse(self):
        items = []
        in_section = False
        collecting = False
        current_lines = []
        script_idx = 0

        def flush():
            nonlocal collecting, current_lines, script_idx
            if collecting and current_lines:
                script_idx += 1
                items.append({
                    "category": "信息转述录音稿",
                    "index": script_idx,
                    "text": sanitize('\n'.join(current_lines)),
                })
            collecting = False
            current_lines = []

        for _, text, _ in self.paras:
            # ---- 章节开始 ----
            if self.RE_SECTION_START.search(text):
                flush()
                in_section = True
                continue

            # ---- 其他「第X节」→ 遇到其他题型的章节，停止采集 ----
            if in_section and self.RE_ANY_SECTION.search(text):
                if not self.RE_SECTION_START.search(text):
                    flush()
                    in_section = False
                    continue

            # ---- 章节结束（第二节 / 参考答案） ----
            if in_section and self.RE_SECTION_END.search(text):
                flush()
                in_section = False
                continue

            if not in_section:
                continue

            # ---- 录音稿标记 ----
            m = self.RE_SCRIPT.match(text)
            if m:
                flush()
                collecting = True
                remainder = m.group(1).strip()
                if remainder:
                    current_lines.append(remainder)
                continue

            # ---- 收集录音稿内容 ----
            if collecting:
                current_lines.append(text)

        flush()
        return self._result(items)


# ============================================================================
# 6. 模仿朗读解析器
# ============================================================================

class ImitationReadingParser(BaseParser):
    """
    解析「模仿朗读」文档。
    提取每个单元的「外网」(通常2篇) 和「教材」(通常1篇) 朗读素材。

    文档结构示例:
        7年级上学期模仿朗读试题
        U5：
          一、模仿朗读（共 6 分）
          听以下短文一遍...
          外网：
          English text 1
          外网：
          English text 2
          教材：
          English text 3
          请听录音。...
        U6：
          ...

    支持同一文件内包含多个单元。
    """

    DOC_TYPE = "模仿朗读"

    # 单元标记：U5 / U5： / u5 / U 5
    RE_UNIT = re.compile(r'^[Uu]\s*(\d+)\s*[：:]?\s*$')
    # 来源标记：外网： / 教材：
    RE_SOURCE = re.compile(r'^(外网|教材)\s*[：:]\s*$')
    # 结束标记
    RE_END = re.compile(r'请听录音|开始录音|停止录音')

    def parse(self):
        items = []
        current_unit = ""
        current_source = None    # "外网" 或 "教材"
        current_lines = []

        def flush():
            nonlocal current_lines, current_source
            if current_source and current_lines:
                items.append({
                    "category": f"模仿朗读-{current_source}",
                    "unit": current_unit,
                    "source": current_source,
                    "text": sanitize('\n'.join(current_lines)),
                })
            current_lines = []
            current_source = None

        for _, text, _ in self.paras:
            # ---- 单元标记 ----
            m_unit = self.RE_UNIT.match(text)
            if m_unit:
                flush()
                current_unit = f"U{m_unit.group(1)}"
                continue

            # ---- 来源标记 ----
            m_src = self.RE_SOURCE.match(text)
            if m_src:
                flush()
                current_source = m_src.group(1)
                continue

            # ---- 结束标记 ----
            if self.RE_END.search(text):
                flush()
                continue

            # ---- 收集朗读内容 ----
            if current_source:
                # 跳过纯中文说明文字
                if not is_chinese(text):
                    current_lines.append(text)

        flush()
        return self._result(items)


# ============================================================================
# 7. 词汇解析器（Excel 单词导入模板）
# ============================================================================

class ExcelVocabularyParser(BaseParser):
    """
    解析 Excel 单词导入模板（.xlsx），提取「单词名称」和「例句」两列。

    每个单词生成两条 TTS 条目：
      1. 单词本身 — 使用默认女声 Amanda，命名「单词1」「单词2」…
      2. 例句     — 使用默认女声 Amanda，命名「句子1」「句子2」…

    Excel 结构（第一行为表头）：
      A: 单元    (如 Unit 6)
      B: 课时信息 (如 Understanding ideas)
      C: 单词名称 (如 pigeon)
      D: 美式音标
      E: 英式音标
      F: 词性
      G: 释义
      H: 例句    (如 A pigeon is standing near the window.)
      I: 翻译

    解析器自动识别「单词名称」和「例句」列的位置（按表头匹配），
    无需硬编码列号。
    """

    DOC_TYPE = "词汇"

    # Excel 文件不使用 load_paragraphs
    _SKIP_LOAD_PARAGRAPHS = True

    # 表头匹配关键词
    WORD_HEADER_KEYWORDS = ("单词名称", "单词")
    SENTENCE_HEADER_KEYWORDS = ("例句",)

    def parse(self):
        if not _OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "解析 Excel 文件需要 openpyxl 库，请运行: pip install openpyxl"
            )

        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        ws = wb.active

        # ---- 识别表头列号 ----
        word_col = None
        sentence_col = None
        headers = {}
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                header = str(cell.value).strip()
                col_idx = cell.column
                headers[col_idx] = header
                if word_col is None:
                    for kw in self.WORD_HEADER_KEYWORDS:
                        if kw in header:
                            word_col = col_idx
                            break
                if sentence_col is None:
                    for kw in self.SENTENCE_HEADER_KEYWORDS:
                        if kw in header:
                            sentence_col = col_idx
                            break

        if word_col is None:
            raise ValueError(
                "未找到「单词名称」列，请确认 Excel 表头包含「单词名称」或「单词」"
            )
        if sentence_col is None:
            raise ValueError(
                "未找到「例句」列，请确认 Excel 表头包含「例句」"
            )

        # ---- 提取数据行 ----
        items = []
        word_seq = 0    # 单词序号
        sentence_seq = 0  # 例句序号

        for row in ws.iter_rows(min_row=2, values_only=True):
            # openpyxl read_only 模式返回的行可能短于总列数
            word_val = row[word_col - 1] if len(row) >= word_col else None
            sentence_val = row[sentence_col - 1] if len(row) >= sentence_col else None

            word_text = sanitize(str(word_val).strip()) if word_val else ""
            sentence_text = sanitize(str(sentence_val).strip()) if sentence_val else ""

            # 跳过空行
            if not word_text and not sentence_text:
                continue

            # 单词条目：使用统一默认女声 Amanda，不设置单词专用音色。
            if word_text:
                word_seq += 1
                items.append({
                    "category": "单词",
                    "number": word_seq,
                    "filename_stem": f"单词{word_seq}",
                    "voice": "female",
                    "text": word_text,
                })

            # 例句条目：同样使用统一默认女声 Amanda。
            if sentence_text:
                sentence_seq += 1
                items.append({
                    "category": "例句",
                    "number": sentence_seq,
                    "filename_stem": f"句子{sentence_seq}",
                    "voice": "female",
                    "text": sentence_text,
                })

        wb.close()
        return self._result(items)


# ============================================================================
# 文档类型自动识别
# ============================================================================

PARSER_MAP = {
    "信息获取": InfoAcquisitionParser,
    "听后选择": ListeningSelectionParser,
    "听后应答": ListeningResponseParser,
    "课文跟读": TextReadingParser,
    "信息转述及询问": InfoRetellingParser,
    "模仿朗读": ImitationReadingParser,
    "词汇": ExcelVocabularyParser,
}


def detect_doc_type(filename):
    """根据文件名自动识别文档类型，返回类型名或 None

    对于 .xlsx 文件，统一归为「词汇」类型。
    """
    # Excel 文件统一归为词汇类型
    if filename.lower().endswith('.xlsx'):
        return "词汇"
    if '信息获取' in filename:
        return "信息获取"
    if '听后选择' in filename:
        return "听后选择"
    if '听后应答' in filename:
        return "听后应答"
    if '课文跟读' in filename:
        return "课文跟读"
    if '信息转述' in filename:
        return "信息转述及询问"
    if '模仿朗读' in filename:
        return "模仿朗读"
    if '词汇' in filename:
        return "词汇"
    return None


# ============================================================================
# 内容自动检测（支持单文档含多题型）
# ============================================================================

# 题型 → 内容标记正则列表（任一匹配即认为文档包含该题型）
CONTENT_MARKERS = {
    "信息获取": [
        re.compile(r'第一节\s*听选信息'),
        re.compile(r'听选信息'),
    ],
    "听后选择": [
        # 与「听选信息」区分，避免旧题型被重复解析；沿用解析器的
        # 行首标题证据，避免普通正文触发新题型。
        ListeningSelectionParser.RE_SECTION_START,
    ],
    "听后应答": [
        ListeningResponseParser.RE_SECTION_START,
        ListeningResponseParser.RE_PROMPT,
    ],
    "课文跟读": [
        re.compile(r'句子跟读'),
        re.compile(r'段落跟读'),
        re.compile(r'语篇跟读'),
    ],
    "信息转述及询问": [
        re.compile(r'第一节\s*信息转述'),
        re.compile(r'信息转述'),
    ],
    "模仿朗读": [
        re.compile(r'模仿朗读'),
        re.compile(r'外网\s*[：:]'),
        re.compile(r'教材\s*[：:]'),
    ],
    "词汇": [
        re.compile(r'词汇例句'),
        re.compile(r'词汇整理'),
        re.compile(r'单词名称'),
        re.compile(r'例句'),
    ],
}


def detect_types_in_content(paras):
    """
    根据文档内容自动识别包含的题型。
    返回检测到的题型名称列表，保持固定顺序。
    """
    full_text = '\n'.join(text for _, text, _ in paras)
    detected = []
    for doc_type in PARSER_MAP:  # 按 PARSER_MAP 的固定顺序
        markers = CONTENT_MARKERS.get(doc_type, [])
        for marker in markers:
            if marker.search(full_text):
                detected.append(doc_type)
                break
    return detected


def parse_document_auto(filepath):
    """
    自动检测文档类型并解析，支持包含多个题型的文档。

    对上传的文档运行所有匹配的解析器，收集非空结果。
    返回 (results_list, summary_str)。

    对于 .xlsx 文件，直接使用 ExcelVocabularyParser 解析。
    """
    filename = os.path.basename(filepath)

    # ---- Excel 文件直接走词汇解析器 ----
    if filename.lower().endswith('.xlsx'):
        doc_type = detect_doc_type(filename)
        if doc_type is None:
            return [], "未识别到任何题型内容"
        parser_cls = PARSER_MAP.get(doc_type)
        if parser_cls is None:
            return [], f"未找到题型 {doc_type} 的解析器"
        try:
            parser = parser_cls(filepath)
            result = parser.parse()
        except Exception as e:
            return [], f"解析失败: {e}"
        if result["item_count"] == 0:
            return [], "未提取到任何内容"
        return [result], f"检测到 1 种题型，成功提取 {result['item_count']} 条内容"

    # ---- Word 文档走原有逻辑 ----
    try:
        paras = load_paragraphs(filepath)
    except Exception as e:
        return [], f"文档加载失败: {e}"

    if not paras:
        return [], "文档内容为空"

    detected_types = detect_types_in_content(paras)

    if not detected_types:
        return [], "未识别到任何题型内容"

    results = []
    errors = []
    for doc_type in detected_types:
        parser_cls = PARSER_MAP.get(doc_type)
        if parser_cls is None:
            continue
        try:
            parser = parser_cls(filepath)
            result = parser.parse()
            if result["item_count"] > 0:
                results.append(result)
        except Exception as e:
            errors.append(f"{doc_type}: {e}")

    total = sum(r["item_count"] for r in results)
    parts = [f"检测到 {len(detected_types)} 种题型"]
    if results:
        parts.append(f"成功提取 {total} 条内容")
    if errors:
        parts.append(f"{len(errors)} 种解析出错")
    summary = "，".join(parts)

    return results, summary


# ============================================================================
# 主函数
# ============================================================================

def main():
    """主函数：遍历 examples/documents，解析所有 Word/Excel 文档。"""
    print("=" * 70)
    print("文档解析脚本")
    print(f"输入目录: {WORD_DIR}")
    print(f"输出目录: {OUTPUT_DIR}")
    print("=" * 70)

    if not os.path.isdir(WORD_DIR):
        print(f"[警告] 输入目录不存在: {WORD_DIR}")
        return

    word_files = [
        f for f in os.listdir(WORD_DIR)
        if (f.endswith('.docx') or f.endswith('.xlsx')) and not f.startswith('~$')
    ]

    if not word_files:
        print("[警告] 输入目录中没有找到 .docx 或 .xlsx 文件")
        return

    all_results = []

    for fname in sorted(word_files):
        filepath = os.path.join(WORD_DIR, fname)
        doc_type = detect_doc_type(fname)

        if doc_type is None:
            print(f"\n[跳过] 未识别类型的文件: {fname}")
            continue

        parser_cls = PARSER_MAP[doc_type]
        try:
            parser = parser_cls(filepath)
            result = parser.parse()
        except Exception as e:
            print(f"\n[错误] 解析失败: {fname} — {e}")
            continue

        all_results.append(result)

        # 打印摘要
        print(f"\n[{doc_type}] {fname}")
        print(f"  共提取 {result['item_count']} 条内容:")

        for item in result["items"]:
            preview = item["text"][:80].replace('\n', ' ')
            cat = item["category"]

            if "number" in item:
                print(f"  · [{cat}] #{item['number']:>2}  {preview}...")
            elif "sentence_number" in item and "discourse_number" in item:
                print(f"  · [{cat}] 语篇{item['discourse_number']}-{item['sentence_number']}  {preview}...")
            elif "sentence_number" in item:
                print(f"  · [{cat}] 句{item['sentence_number']}  {preview}...")
            elif "index" in item:
                print(f"  · [{cat}] #{item['index']:>2}  {preview}...")
            elif "unit" in item:
                print(f"  · [{cat}] ({item['unit']:<3}) {preview}...")
            elif "discourse_number" in item:
                print(f"  · [{cat}] 语篇{item['discourse_number']}  {preview}...")
            else:
                print(f"  · [{cat}] {preview}...")

    # 保存汇总 JSON
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, "parsed_results.json")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)

    print("\n" + "=" * 70)
    total_items = sum(r["item_count"] for r in all_results)
    print(f"解析完成！共处理 {len(all_results)} 个文件，提取 {total_items} 条内容")
    print(f"结果已保存到: {output_path}")
    print("=" * 70)


if __name__ == "__main__":
    main()
